#!/usr/bin/env python3
# -*- coding: utf8 -*-


from openpyxl import Workbook
import sqlite3
import datetime

SQLFile = 'data/taizhang.' + datetime.datetime.now().strftime('%Y%m%d') +'.sqlite3'
database = sqlite3.connect(SQLFile)
WorkBookFile = 'data/taizhang.' + datetime.datetime.now().strftime('%Y%m%d') +'.xlsx'
WorkBook = Workbook()


def export_mode(WorkSheet1, query):
    cursor = database.cursor()
    cursor.execute(query)
    results = cursor.fetchall()
    i = 2
    for result in results:
        j = 1
        for r in result:
            # print j
            WorkSheet1.cell(row=i, column=j).value = r
            j = j + 1
        i = i + 1
    cursor.close()


def export_guestos(WorkBook, sheetname):
    # WorkBook.remove_sheet(WorkBook.get_sheet_by_name(sheetname))
    WorkSheet1 = WorkBook.create_sheet(title=sheetname)

    # 导入标题栏
    Title = "status>>>inventory>>>hostname>>>os>>>proiplist>>maniplist>>>othiplist>>>diskuse>>>disksize>>>vc>>>cluster>>>esxhost>>>vmx>>>mem>>>cpu>>>ethernet>>>abbsys>>>uuid>>>mark"
    Title = Title.split('>>>')
    for col in range(0, len(Title)):
        WorkSheet1.cell(column=col+1, row=1).value = Title[col]

    query = "select * from guestos"
    export_mode(WorkSheet1, query)


def export_esxhost(WorkBook, sheetname):

    worksheet1 = WorkBook.create_sheet(title=sheetname)

    # 导入标题栏
    Title = "VC>>>主机>>>集群>>>PC型号>>>CPU型号>>>CPU总core>>>CPU已分配core>>>总内存/G>>>内存已分配/G"
    Title = Title.split('>>>')
    for col in range(0, len(Title)):
        worksheet1.cell(column=col+1, row=1).value = Title[col]

    query = "select * from vt_host"
    export_mode(worksheet1, query)


def main():
    # WorkBook = load_workbook(filename, data_only=True)

    sheetname = 'GuestOS'
    export_guestos(WorkBook, sheetname)

    sheetname = 'esxHost'
    export_esxhost(WorkBook, sheetname)

    WorkBook.save(filename=WorkBookFile)

    # Close the database connection
    database.close()


if __name__ == "__main__":
    main()
